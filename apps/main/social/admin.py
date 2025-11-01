from django.contrib import admin
from django.utils.translation import gettext_lazy as _
from django.utils.html import format_html
from django.contrib import messages
from django.shortcuts import redirect
from .models import (
    Post, Comment, Like, Follow, UserProfile, 
    Share, Hashtag, PostHashtag, CommentLike,
    Report, ModerationAction, ContentFilter, ModerationLog, ReportFilterFlag,
    VerificationRequest
)
from core.admin import BaseModelAdmin
from django.utils import timezone


@admin.register(Post)
class PostAdmin(BaseModelAdmin):
    list_display = [
        'author', 'content_preview', 'is_public', 'is_pinned', 
        'likes_count', 'comments_count', 'views_count', 'shares_count', 'created_at'
    ]
    list_filter = [
        'is_public', 'is_pinned', 'is_edited', 'created_at', 'author'
    ]
    search_fields = [
        'content', 'author__username', 'author__first_name', 'author__last_name'
    ]
    readonly_fields = [
        'likes_count', 'comments_count', 'views_count', 'shares_count', 
        'created_at', 'updated_at', 'edited_at', 'engagement_rate'
    ]
    date_hierarchy = 'created_at'
    fieldsets = (
        (_('Informações Básicas'), {
            'fields': ('author', 'content', 'is_public', 'is_pinned')
        }),
        (_('Mídia'), {
            'fields': ('image', 'video', 'link', 'link_title', 'link_description', 'link_image'),
            'classes': ('collapse',)
        }),
        (_('Estatísticas'), {
            'fields': ('likes_count', 'comments_count', 'views_count', 'shares_count', 'engagement_rate'),
            'classes': ('collapse',)
        }),
        (_('Metadados'), {
            'fields': ('created_at', 'updated_at', 'is_edited', 'edited_at'),
            'classes': ('collapse',)
        }),
    )
    
    def content_preview(self, obj):
        return obj.content[:100] + '...' if len(obj.content) > 100 else obj.content
    content_preview.short_description = _('Conteúdo')
    
    def engagement_rate(self, obj):
        return f"{obj.engagement_rate:.1f}%"
    engagement_rate.short_description = _('Taxa de Engajamento')


@admin.register(Comment)
class CommentAdmin(BaseModelAdmin):
    list_display = [
        'author', 'post', 'content_preview', 'parent', 
        'likes_count', 'is_edited', 'created_at'
    ]
    list_filter = [
        'created_at', 'is_edited', 'author', 'post'
    ]
    search_fields = [
        'content', 'author__username', 'post__content'
    ]
    readonly_fields = [
        'likes_count', 'created_at', 'is_edited', 'edited_at'
    ]
    date_hierarchy = 'created_at'
    
    def content_preview(self, obj):
        return obj.content[:50] + '...' if len(obj.content) > 50 else obj.content
    content_preview.short_description = _('Conteúdo')


@admin.register(CommentLike)
class CommentLikeAdmin(BaseModelAdmin):
    list_display = ['user', 'comment', 'created_at']
    list_filter = ['created_at', 'user', 'comment__post']
    search_fields = [
        'user__username', 'comment__content', 'comment__post__content'
    ]
    readonly_fields = ['created_at']
    date_hierarchy = 'created_at'


@admin.register(Like)
class LikeAdmin(BaseModelAdmin):
    list_display = ['user', 'post', 'reaction_type', 'created_at']
    list_filter = [
        'reaction_type', 'created_at', 'user', 'post'
    ]
    search_fields = [
        'user__username', 'post__content'
    ]
    readonly_fields = ['created_at']
    date_hierarchy = 'created_at'


@admin.register(Share)
class ShareAdmin(BaseModelAdmin):
    list_display = [
        'user', 'original_post', 'comment_preview', 'created_at'
    ]
    list_filter = ['created_at', 'user', 'original_post__author']
    search_fields = [
        'user__username', 'original_post__content', 'comment'
    ]
    readonly_fields = ['created_at']
    date_hierarchy = 'created_at'
    
    def comment_preview(self, obj):
        return obj.comment[:50] + '...' if obj.comment and len(obj.comment) > 50 else obj.comment
    comment_preview.short_description = _('Comentário')


@admin.register(Follow)
class FollowAdmin(BaseModelAdmin):
    list_display = [
        'follower', 'following', 'notifications_enabled', 'created_at'
    ]
    list_filter = [
        'notifications_enabled', 'created_at', 'follower', 'following'
    ]
    search_fields = [
        'follower__username', 'following__username'
    ]
    readonly_fields = ['created_at']
    date_hierarchy = 'created_at'


@admin.register(UserProfile)
class UserProfileAdmin(BaseModelAdmin):
    list_display = [
        'user', 'bio_preview', 'website_preview', 'is_private', 'location', 
        'total_posts', 'total_likes_received', 'created_at'
    ]
    list_filter = [
        'is_private', 'show_email', 'show_phone', 'allow_messages', 
        'gender', 'location', 'created_at'
    ]
    search_fields = [
        'user__username', 'user__first_name', 'user__last_name', 
        'bio', 'location', 'interests', 'website'
    ]
    readonly_fields = [
        'total_posts', 'total_likes_received', 'total_comments_received', 'created_at', 'updated_at'
    ]
    fieldsets = (
        (_('Informações Básicas'), {
            'fields': ('user', 'bio', 'location', 'birth_date', 'gender', 'phone')
        }),
        (_('Mídia'), {
            'fields': ('avatar', 'cover_image'),
            'classes': ('collapse',)
        }),
        (_('Links'), {
            'fields': ('website', 'social_links'),
            'classes': ('collapse',)
        }),
        (_('Interesses'), {
            'fields': ('interests',),
            'classes': ('collapse',)
        }),
        (_('Privacidade'), {
            'fields': ('is_private', 'show_email', 'show_phone', 'allow_messages')
        }),
        (_('Estatísticas'), {
            'fields': ('total_posts', 'total_likes_received', 'total_comments_received'),
            'classes': ('collapse',)
        }),
    )
    
    def bio_preview(self, obj):
        return obj.bio[:50] + '...' if obj.bio and len(obj.bio) > 50 else obj.bio
    bio_preview.short_description = _('Biografia')

    def website_preview(self, obj):
        if obj.website:
            # Remove protocolo para exibição mais limpa
            clean_url = obj.website.replace('https://', '').replace('http://', '').replace('www.', '')
            # Trunca se muito longo
            display_url = clean_url[:30] + '...' if len(clean_url) > 30 else clean_url
            return format_html('<a href="{}" target="_blank" title="{}">{}</a>', 
                             obj.website, obj.website, display_url)
        return '-'
    website_preview.short_description = _('Website')
    website_preview.admin_order_field = 'website'


@admin.register(Hashtag)
class HashtagAdmin(BaseModelAdmin):
    list_display = ['name', 'posts_count', 'created_at']
    list_filter = ['created_at']
    search_fields = ['name', 'description']
    readonly_fields = ['posts_count', 'created_at', 'updated_at']
    ordering = ['-posts_count', '-created_at']
    
    fieldsets = (
        (_('Informações Básicas'), {
            'fields': ('name', 'description')
        }),
        (_('Estatísticas'), {
            'fields': ('posts_count', 'created_at', 'updated_at'),
            'classes': ('collapse',)
        }),
    )


@admin.register(PostHashtag)
class PostHashtagAdmin(BaseModelAdmin):
    list_display = ['post', 'hashtag', 'created_at']
    list_filter = ['hashtag', 'created_at']
    search_fields = ['post__content', 'hashtag__name']
    readonly_fields = ['created_at', 'updated_at']
    ordering = ['-created_at']


# ============================================================================
# ADMIN DE MODERAÇÃO
# ============================================================================

class ReportFilterFlagInline(admin.TabularInline):
    model = ReportFilterFlag
    extra = 0
    readonly_fields = ['content_filter', 'matched_pattern', 'confidence_score', 'created_at']
    can_delete = False


@admin.register(Report)
class ReportAdmin(BaseModelAdmin):
    list_display = [
        'get_reported_content_short', 'report_type', 'reporter', 'status', 
        'priority', 'get_filters_count', 'assigned_moderator', 'created_at'
    ]
    list_filter = [
        'report_type', 'status', 'priority', 'created_at', 'assigned_moderator',
        'filter_flags__content_filter'
    ]
    search_fields = [
        'description', 'reporter__username', 'reported_post__content',
        'reported_comment__content', 'reported_user__username',
        'filter_flags__content_filter__name'
    ]
    readonly_fields = [
        'similar_reports_count', 'created_at', 'updated_at', 'resolved_at',
        'get_filter_flags_display'
    ]
    date_hierarchy = 'created_at'
    ordering = ['-priority', '-created_at']
    inlines = [ReportFilterFlagInline]
    
    fieldsets = (
        (_('Informações da Denúncia'), {
            'fields': ('reporter', 'report_type', 'description', 'status', 'priority')
        }),
        (_('Conteúdo Reportado'), {
            'fields': ('reported_post', 'reported_comment', 'reported_user'),
            'classes': ('collapse',)
        }),
        (_('Moderação'), {
            'fields': ('assigned_moderator', 'moderator_notes', 'resolved_at')
        }),
        (_('Estatísticas'), {
            'fields': ('similar_reports_count', 'created_at', 'updated_at'),
            'classes': ('collapse',)
        }),
    )
    
    actions = ['assign_to_moderator', 'mark_as_resolved', 'mark_as_dismissed']
    
    def get_reported_content_short(self, obj):
        content = obj.get_reported_content()
        return content[:50] + '...' if len(content) > 50 else content
    get_reported_content_short.short_description = _('Conteúdo Reportado')
    
    def get_filters_count(self, obj):
        count = obj.filter_flags.count()
        if count == 0:
            return _('Manual')
        elif count == 1:
            filter_name = obj.filter_flags.first().content_filter.name
            return format_html(
                '<span style="background: #e3f2fd; padding: 2px 6px; border-radius: 3px; font-size: 11px;">{}</span>',
                filter_name[:15] + '...' if len(filter_name) > 15 else filter_name
            )
        else:
            return format_html(
                '<span style="background: #fff3e0; padding: 2px 6px; border-radius: 3px; font-size: 11px; color: #f57c00;">{} filtros</span>',
                count
            )
    get_filters_count.short_description = _('Filtros')
    
    def get_filter_flags_display(self, obj):
        flags = obj.filter_flags.select_related('content_filter').all()
        if not flags:
            return _('Denúncia manual - não foi gerada por filtros automáticos')
        
        html_parts = []
        for flag in flags:
            confidence_color = '#4caf50' if flag.confidence_score >= 0.8 else '#ff9800' if flag.confidence_score >= 0.5 else '#f44336'
            html_parts.append(format_html(
                '<div style="margin: 8px 0; padding: 8px; border: 1px solid #ddd; border-radius: 4px;">'
                '<strong>{}</strong> <span style="color: {};">({:.0%})</span><br>'
                '<small style="color: #666;">{}</small><br>'
                '<code style="background: #f5f5f5; padding: 2px 4px; border-radius: 2px; font-size: 11px;">{}</code>'
                '</div>',
                flag.content_filter.name,
                confidence_color,
                flag.confidence_score,
                flag.content_filter.get_filter_type_display(),
                flag.matched_pattern[:100] + '...' if flag.matched_pattern and len(flag.matched_pattern) > 100 else flag.matched_pattern or 'N/A'
            ))
        return format_html(''.join(html_parts))
    get_filter_flags_display.short_description = _('Detalhes dos Filtros Acionados')
    
    def assign_to_moderator(self, request, queryset):
        """Ação para atribuir denúncias a um moderador"""
        from django.contrib.auth import get_user_model
        User = get_user_model()
        
        # Buscar moderadores disponíveis
        moderators = User.objects.filter(is_staff=True)
        
        # Atribuir denúncias pendentes
        pending_reports = queryset.filter(status='pending')
        for i, report in enumerate(pending_reports):
            moderator = moderators[i % len(moderators)]
            report.assigned_moderator = moderator
            report.status = 'reviewing'
            report.save()
        
        self.message_user(
            request, 
            f'{pending_reports.count()} denúncias foram atribuídas a moderadores.'
        )
    assign_to_moderator.short_description = _('Atribuir a moderadores')
    
    def mark_as_resolved(self, request, queryset):
        """Marcar denúncias como resolvidas"""
        queryset.update(status='resolved', resolved_at=timezone.now())
        self.message_user(request, f'{queryset.count()} denúncias foram marcadas como resolvidas.')
    mark_as_resolved.short_description = _('Marcar como resolvidas')
    
    def mark_as_dismissed(self, request, queryset):
        """Marcar denúncias como descartadas"""
        queryset.update(status='dismissed')
        self.message_user(request, f'{queryset.count()} denúncias foram descartadas.')
    mark_as_dismissed.short_description = _('Descartar denúncias')


@admin.register(ModerationAction)
class ModerationActionAdmin(BaseModelAdmin):
    list_display = [
        'action_type', 'get_target_description_short', 'moderator', 
        'is_active', 'created_at'
    ]
    list_filter = [
        'action_type', 'is_active', 'suspension_type', 'created_at', 'moderator'
    ]
    search_fields = [
        'reason', 'moderator__username', 'target_post__content',
        'target_comment__content', 'target_user__username'
    ]
    readonly_fields = [
        'suspension_end_date', 'expires_at', 'created_at', 'updated_at'
    ]
    date_hierarchy = 'created_at'
    ordering = ['-created_at']
    
    fieldsets = (
        (_('Informações da Ação'), {
            'fields': ('moderator', 'action_type', 'reason')
        }),
        (_('Alvo da Ação'), {
            'fields': ('target_post', 'target_comment', 'target_user'),
            'classes': ('collapse',)
        }),
        (_('Suspensão'), {
            'fields': ('suspension_duration', 'suspension_type', 'suspension_end_date'),
            'classes': ('collapse',)
        }),
        (_('Controle'), {
            'fields': ('is_active', 'expires_at', 'notify_user', 'notification_message')
        }),
        (_('Metadados'), {
            'fields': ('created_at', 'updated_at'),
            'classes': ('collapse',)
        }),
    )
    
    actions = ['deactivate_actions', 'extend_suspension']
    
    def get_target_description_short(self, obj):
        target = obj.get_target_description()
        return target[:50] + '...' if len(target) > 50 else target
    get_target_description_short.short_description = _('Alvo')
    
    def deactivate_actions(self, request, queryset):
        """Desativar ações de moderação"""
        queryset.update(is_active=False)
        self.message_user(request, f'{queryset.count()} ações foram desativadas.')
    deactivate_actions.short_description = _('Desativar ações')
    
    def extend_suspension(self, request, queryset):
        """Estender suspensões"""
        from datetime import timedelta
        
        suspended_actions = queryset.filter(
            action_type__in=['suspend_user', 'restrict_user'],
            is_active=True
        )
        
        for action in suspended_actions:
            if action.suspension_end_date:
                action.suspension_end_date += timedelta(days=7)
                action.expires_at = action.suspension_end_date
                action.save()
        
        self.message_user(
            request, 
            f'Suspensões de {suspended_actions.count()} ações foram estendidas por 7 dias.'
        )
    extend_suspension.short_description = _('Estender suspensões por 7 dias')


@admin.register(ContentFilter)
class ContentFilterAdmin(BaseModelAdmin):
    list_display = [
        'name', 'filter_type', 'action', 'is_active', 'matches_count', 'last_matched'
    ]
    list_filter = [
        'filter_type', 'action', 'is_active', 'case_sensitive', 'created_at'
    ]
    search_fields = ['name', 'pattern', 'description']
    readonly_fields = [
        'matches_count', 'last_matched', 'created_at', 'updated_at'
    ]
    ordering = ['-is_active', 'name']
    
    fieldsets = (
        (_('Informações do Filtro'), {
            'fields': ('name', 'filter_type', 'pattern', 'action', 'description')
        }),
        (_('Configurações'), {
            'fields': ('is_active', 'case_sensitive')
        }),
        (_('Aplicação'), {
            'fields': ('apply_to_posts', 'apply_to_comments', 'apply_to_usernames')
        }),
        (_('Estatísticas'), {
            'fields': ('matches_count', 'last_matched', 'created_at', 'updated_at'),
            'classes': ('collapse',)
        }),
    )
    
    actions = ['activate_filters', 'deactivate_filters', 'reset_statistics']
    
    def activate_filters(self, request, queryset):
        """Ativar filtros selecionados"""
        queryset.update(is_active=True)
        self.message_user(request, f'{queryset.count()} filtros foram ativados.')
    activate_filters.short_description = _('Ativar filtros')
    
    def deactivate_filters(self, request, queryset):
        """Desativar filtros selecionados"""
        queryset.update(is_active=False)
        self.message_user(request, f'{queryset.count()} filtros foram desativados.')
    deactivate_filters.short_description = _('Desativar filtros')
    
    def reset_statistics(self, request, queryset):
        """Resetar estatísticas dos filtros"""
        queryset.update(matches_count=0, last_matched=None)
        self.message_user(request, f'Estatísticas de {queryset.count()} filtros foram resetadas.')
    reset_statistics.short_description = _('Resetar estatísticas')


@admin.register(ReportFilterFlag)
class ReportFilterFlagAdmin(BaseModelAdmin):
    list_display = [
        'report', 'content_filter', 'get_matched_pattern_short', 'confidence_score', 'created_at'
    ]
    list_filter = [
        'content_filter', 'confidence_score', 'created_at'
    ]
    search_fields = [
        'report__description', 'content_filter__name', 'matched_pattern'
    ]
    readonly_fields = ['created_at', 'updated_at']
    date_hierarchy = 'created_at'
    ordering = ['-created_at']
    
    def get_matched_pattern_short(self, obj):
        if not obj.matched_pattern:
            return '-'
        pattern = obj.matched_pattern
        return pattern[:50] + '...' if len(pattern) > 50 else pattern
    get_matched_pattern_short.short_description = _('Padrão Detectado')


@admin.register(ModerationLog)
class ModerationLogAdmin(BaseModelAdmin):
    list_display = [
        'action_type', 'moderator', 'target_type', 'target_id', 'created_at'
    ]
    list_filter = [
        'action_type', 'target_type', 'created_at', 'moderator'
    ]
    search_fields = [
        'description', 'details', 'moderator__username'
    ]
    readonly_fields = [
        'ip_address', 'user_agent', 'created_at', 'updated_at'
    ]
    date_hierarchy = 'created_at'
    ordering = ['-created_at']
    
    fieldsets = (
        (_('Informações da Ação'), {
            'fields': ('moderator', 'action_type', 'target_type', 'target_id')
        }),
        (_('Detalhes'), {
            'fields': ('description', 'details')
        }),
        (_('Contexto'), {
            'fields': ('ip_address', 'user_agent'),
            'classes': ('collapse',)
        }),
        (_('Metadados'), {
            'fields': ('created_at', 'updated_at'),
            'classes': ('collapse',)
        }),
    )
    
    def has_add_permission(self, request):
        """Impedir criação manual de logs"""
        return False
    
    def has_change_permission(self, request, obj=None):
        """Impedir edição de logs"""
        return False
    

    
    def export_logs_csv(self, request, queryset):
        """Exportar logs selecionados em CSV"""
        import csv
        from django.http import HttpResponse
        
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="moderation_logs.csv"'
        
        writer = csv.writer(response)
        writer.writerow([
            'Data', 'Moderador', 'Tipo de Ação', 'Tipo do Alvo', 'ID do Alvo',
            'Descrição', 'Detalhes', 'IP', 'User Agent'
        ])
        
        for log in queryset:
            writer.writerow([
                log.created_at.strftime('%Y-%m-%d %H:%M:%S'),
                log.moderator.username if log.moderator else 'Sistema',
                log.get_action_type_display(),
                log.target_type,
                log.target_id,
                log.description,
                log.details or '',
                log.ip_address or '',
                log.user_agent or ''
            ])
        
        return response
    export_logs_csv.short_description = _('Exportar logs CSV')

    def export_logs_excel(self, request, queryset):
        """Exportar logs selecionados em Excel formatado"""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            messages.error(request, _('openpyxl não está instalado. Use: pip install openpyxl'))
            return redirect(request.META.get('HTTP_REFERER', '/admin/'))
        
        from django.http import HttpResponse
        
        # Criar workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Logs de Moderação"
        
        # Estilos
        header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
        
        border_thin = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        data_font = Font(name='Arial', size=10)
        data_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        
        # Cabeçalhos
        headers = [
            'Data/Hora', 'Moderador', 'Tipo de Ação', 'Tipo do Alvo', 
            'ID do Alvo', 'Descrição', 'Detalhes', 'Endereço IP', 'Navegador'
        ]
        
        # Configurar cabeçalho
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border_thin
        
        # Dados
        for row_num, log in enumerate(queryset, 2):
            # Data formatada
            ws.cell(row=row_num, column=1, value=log.created_at.strftime('%d/%m/%Y %H:%M:%S'))
            
            # Moderador
            moderator_name = log.moderator.get_full_name() if log.moderator and log.moderator.get_full_name() else (
                log.moderator.username if log.moderator else 'Sistema Automático'
            )
            ws.cell(row=row_num, column=2, value=moderator_name)
            
            # Tipo de ação
            ws.cell(row=row_num, column=3, value=log.get_action_type_display())
            
            # Tipo do alvo
            ws.cell(row=row_num, column=4, value=log.target_type.title())
            
            # ID do alvo
            ws.cell(row=row_num, column=5, value=log.target_id)
            
            # Descrição
            ws.cell(row=row_num, column=6, value=log.description[:500])  # Limitar tamanho
            
            # Detalhes
            details = log.details[:500] if log.details else 'N/A'
            ws.cell(row=row_num, column=7, value=details)
            
            # IP
            ws.cell(row=row_num, column=8, value=log.ip_address or 'N/A')
            
            # User Agent simplificado
            user_agent = 'N/A'
            if log.user_agent:
                if 'Chrome' in log.user_agent:
                    user_agent = 'Chrome'
                elif 'Firefox' in log.user_agent:
                    user_agent = 'Firefox'
                elif 'Safari' in log.user_agent:
                    user_agent = 'Safari'
                elif 'Edge' in log.user_agent:
                    user_agent = 'Edge'
                else:
                    user_agent = 'Outro'
            ws.cell(row=row_num, column=9, value=user_agent)
            
            # Aplicar estilos às células de dados
            for col_num in range(1, 10):
                cell = ws.cell(row=row_num, column=col_num)
                cell.font = data_font
                cell.alignment = data_alignment
                cell.border = border_thin
        
        # Ajustar largura das colunas
        column_widths = {
            1: 18,  # Data/Hora
            2: 20,  # Moderador
            3: 25,  # Tipo de Ação
            4: 15,  # Tipo do Alvo
            5: 10,  # ID do Alvo
            6: 40,  # Descrição
            7: 30,  # Detalhes
            8: 15,  # IP
            9: 12,  # Navegador
        }
        
        for col_num, width in column_widths.items():
            ws.column_dimensions[get_column_letter(col_num)].width = width
        
        # Adicionar aba de estatísticas
        stats_ws = wb.create_sheet(title="Estatísticas")
        
        # Cabeçalho da aba de estatísticas
        stats_ws.cell(row=1, column=1, value="RELATÓRIO DE LOGS DE MODERAÇÃO").font = Font(size=16, bold=True)
        stats_ws.cell(row=2, column=1, value=f"Gerado em: {timezone.now().strftime('%d/%m/%Y %H:%M:%S')}")
        stats_ws.cell(row=3, column=1, value=f"Total de registros: {queryset.count()}")
        
        # Estatísticas por tipo de ação
        stats_ws.cell(row=5, column=1, value="Ações por Tipo:").font = Font(bold=True)
        action_stats = {}
        for log in queryset:
            action_type = log.get_action_type_display()
            action_stats[action_type] = action_stats.get(action_type, 0) + 1
        
        row = 6
        for action, count in action_stats.items():
            stats_ws.cell(row=row, column=1, value=action)
            stats_ws.cell(row=row, column=2, value=count)
            row += 1
        
        # Estatísticas por moderador
        stats_ws.cell(row=row + 1, column=1, value="Ações por Moderador:").font = Font(bold=True)
        moderator_stats = {}
        for log in queryset:
            moderator_name = log.moderator.get_full_name() if log.moderator and log.moderator.get_full_name() else (
                log.moderator.username if log.moderator else 'Sistema Automático'
            )
            moderator_stats[moderator_name] = moderator_stats.get(moderator_name, 0) + 1
        
        row += 2
        for moderator, count in moderator_stats.items():
            stats_ws.cell(row=row, column=1, value=moderator)
            stats_ws.cell(row=row, column=2, value=count)
            row += 1
        
        # Ajustar largura das colunas da aba de estatísticas
        stats_ws.column_dimensions['A'].width = 30
        stats_ws.column_dimensions['B'].width = 10
        
        # Preparar resposta
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"logs_moderacao_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        # Salvar workbook na resposta
        wb.save(response)
        return response
    
    export_logs_excel.short_description = _('Exportar logs Excel')

    actions = ['export_logs_csv', 'export_logs_excel']


@admin.register(VerificationRequest)
class VerificationRequestAdmin(BaseModelAdmin):
    """Admin para solicitações de verificação"""
    
    list_display = [
        'user', 'full_name', 'status', 'created_at', 'reviewed_by', 'reviewed_at'
    ]
    
    list_filter = [
        'status', 'created_at', 'reviewed_at', 'is_cpf_locked'
    ]
    
    search_fields = [
        'user__username', 'user__email', 'full_name', 'cpf'
    ]
    
    readonly_fields = [
        'user', 'created_at', 'is_cpf_locked'
    ]
    
    fieldsets = (
        ('Informações do Usuário', {
            'fields': ('user', 'created_at')
        }),
        ('Informações Pessoais', {
            'fields': ('full_name', 'cpf', 'birth_date', 'phone_number')
        }),
        ('Documentação', {
            'fields': ('identity_document',)
        }),
        ('Solicitação', {
            'fields': ('reason',)
        }),
        ('Revisão', {
            'fields': ('status', 'rejection_reason', 'admin_notes', 'reviewed_by', 'reviewed_at')
        }),
        ('Controle', {
            'fields': ('is_cpf_locked',)
        }),
    )
    
    def get_queryset(self, request):
        return super().get_queryset(request).select_related('user', 'reviewed_by')
    
    def save_model(self, request, obj, form, change):
        if not change:  # Nova solicitação
            obj.user = request.user
        else:  # Editando
            if 'status' in form.changed_data:
                obj.reviewed_by = request.user
                obj.reviewed_at = timezone.now()
        
        # O método save() do modelo já cuida de marcar o usuário como verificado
        super().save_model(request, obj, form, change)
    
    def has_add_permission(self, request):
        return False  # Não permitir criar solicitações manualmente
    
    def has_delete_permission(self, request, obj=None):
        return request.user.is_superuser  # Apenas superusuários podem deletar
    
    actions = ['approve_selected', 'reject_selected']
    
    def approve_selected(self, request, queryset):
        """Aprovar solicitações selecionadas"""
        count = 0
        for verification_request in queryset.filter(status='pending'):
            verification_request.status = 'approved'
            verification_request.reviewed_by = request.user
            verification_request.reviewed_at = timezone.now()
            verification_request.save()  # O método save() do modelo já cuida de marcar o usuário como verificado
            count += 1
        
        self.message_user(request, f'{count} solicitação(ões) aprovada(s) com sucesso.')
    
    approve_selected.short_description = "Aprovar solicitações selecionadas"
    
    def reject_selected(self, request, queryset):
        """Rejeitar solicitações selecionadas"""
        count = queryset.filter(status='pending').update(
            status='rejected',
            reviewed_by=request.user,
            reviewed_at=timezone.now()
        )
        self.message_user(request, f'{count} solicitação(ões) rejeitada(s) com sucesso.')
    
    reject_selected.short_description = "Rejeitar solicitações selecionadas"


# Configurações do admin
admin.site.site_header = _('Administração da Rede Social')
admin.site.site_title = _('Rede Social Admin')
admin.site.index_title = _('Gerenciamento da Rede Social')
